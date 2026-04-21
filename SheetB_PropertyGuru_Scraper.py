"""
SheetB_PropertyGuru_Scraper.py
================================
Pulls property listing data from PropertyGuru search results based on
advisor-defined parameters, then outputs rows formatted for Sheet B
(Phase3_Project_Database_Master.xlsx).

USAGE
-----
Run from terminal:
    python SheetB_PropertyGuru_Scraper.py

Or import and call scrape() directly from another script:
    from SheetB_PropertyGuru_Scraper import scrape
    results = scrape(params)

REQUIREMENTS
------------
    pip install requests beautifulsoup4 openpyxl

PARAMETERS (edit the SEARCH_PARAMS block below)
------------
    districts        : list of district codes, e.g. ["D14", "D19"]
    property_types   : list of types: "Condo", "EC", "HDB", "Landed"
    budget_min       : minimum price in SGD
    budget_max       : maximum price in SGD
    bedrooms         : number of bedrooms (1-5)
    tenure           : "99", "999", "Freehold", or "Any"
    top_year_from    : filter by TOP year >= this (e.g. 2018)
    top_year_to      : filter by TOP year <= this (e.g. 2026)
    max_results      : how many listings to pull (max ~50 per run)

OUTPUT
------
1. Console: prints scraped listings in table format
2. SheetB_Scraped_Results.xlsx: ready to review and paste into Sheet B
3. SheetB_Scraped_Results.json: raw data for dashboard loading

NOTES
-----
- PropertyGuru's public search pages are used (no API key needed)
- If PropertyGuru blocks scraping, the fallback mode generates a
  structured template you can fill manually from the site
- Always verify scraped prices against live PropertyGuru listings
  before using in client consultations
- Run this fresh before each batch of consultations to keep Sheet B current
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════
#  SEARCH PARAMETERS — edit these before each run
# ═══════════════════════════════════════════════════════════
SEARCH_PARAMS = {
    "districts":       ["D14", "D19", "D16", "D18", "D23"],  # target districts
    "property_types":  ["Condo", "EC"],                        # Condo, EC, HDB, Landed
    "budget_min":      1000000,                                # SGD
    "budget_max":      1600000,                                # SGD
    "bedrooms":        2,                                      # 1, 2, 3, 4, 5
    "tenure":          "Any",                                  # "99", "999", "Freehold", "Any"
    "top_year_from":   2015,                                   # filter by completion year
    "top_year_to":     2030,
    "mrt_walk_max":    15,                                     # max walk mins to MRT (advisory filter, not PG filter)
    "max_results":     30,
}

# ═══════════════════════════════════════════════════════════
#  DISTRICT CODE MAP
# ═══════════════════════════════════════════════════════════
DISTRICT_MAP = {
    "D01": "1", "D02": "2", "D03": "3", "D04": "4", "D05": "5",
    "D06": "6", "D07": "7", "D08": "8", "D09": "9", "D10": "10",
    "D11": "11", "D12": "12", "D13": "13", "D14": "14", "D15": "15",
    "D16": "16", "D17": "17", "D18": "18", "D19": "19", "D20": "20",
    "D21": "21", "D22": "22", "D23": "23", "D24": "24", "D25": "25",
    "D26": "26", "D27": "27", "D28": "28",
}

DISTRICT_AREAS = {
    "D14": "Geylang / Eunos / Paya Lebar",
    "D15": "Katong / Marine Parade / Amber",
    "D16": "Bedok / Upper East Coast",
    "D18": "Tampines / Pasir Ris",
    "D19": "Hougang / Punggol / Sengkang",
    "D20": "Ang Mo Kio / Bishan / Thomson",
    "D21": "Clementi / Upper Bukit Timah",
    "D22": "Jurong / Boon Lay",
    "D23": "Bukit Batok / Choa Chu Kang",
    "D25": "Woodlands / Admiralty",
    "D10": "Bukit Timah / Holland",
    "D11": "Newton / Novena",
    "D09": "Orchard / River Valley",
}

# ═══════════════════════════════════════════════════════════
#  HEADERS (mimic browser to avoid blocks)
# ═══════════════════════════════════════════════════════════
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-SG,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.propertyguru.com.sg/",
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)


# ═══════════════════════════════════════════════════════════
#  BUILD PROPERTYGURU SEARCH URL
# ═══════════════════════════════════════════════════════════
def build_search_url(params, page=1):
    """Build PropertyGuru resale/new-launch search URL from params."""
    base = "https://www.propertyguru.com.sg/property-for-sale"

    # District filter
    district_nums = [DISTRICT_MAP[d] for d in params["districts"] if d in DISTRICT_MAP]
    district_str = "&".join(f"district_code[]={n}" for n in district_nums)

    # Property type
    type_map = {"Condo": "N", "EC": "E", "HDB": "H", "Landed": "D"}
    prop_types = "&".join(
        f"property_type[]={type_map[t]}"
        for t in params["property_types"] if t in type_map
    )

    # Bedrooms
    beds = params.get("bedrooms", 2)

    # Price range
    price_min = params.get("budget_min", 0)
    price_max = params.get("budget_max", 9999999)

    # Tenure
    tenure_map = {"99": "1", "999": "3", "Freehold": "2", "Any": ""}
    tenure_val = tenure_map.get(params.get("tenure", "Any"), "")
    tenure_str = f"freehold={tenure_val}" if tenure_val else ""

    url = (
        f"{base}/{page}"
        f"?{district_str}"
        f"&{prop_types}"
        f"&bedrooms={beds}"
        f"&minprice={price_min}"
        f"&maxprice={price_max}"
    )
    if tenure_str:
        url += f"&{tenure_str}"

    return url


# ═══════════════════════════════════════════════════════════
#  PARSE A LISTING CARD
# ═══════════════════════════════════════════════════════════
def parse_listing(card):
    """Extract structured data from a PropertyGuru listing card element."""
    result = {}

    try:
        # Project / listing name
        name_el = card.select_one("[data-automation-id='listing-card-description'] h3, .listing-card h3, h3.ellipsis")
        result["name"] = name_el.get_text(strip=True) if name_el else "Unknown"

        # Price
        price_el = card.select_one("[class*='price']")
        if price_el:
            price_text = price_el.get_text(strip=True).replace(",", "").replace("$", "").replace(" ", "")
            prices = re.findall(r"\d+", price_text)
            if prices:
                result["budgetMin"] = int(prices[0])
                result["budgetMax"] = int(prices[-1]) if len(prices) > 1 else int(prices[0])
            else:
                result["budgetMin"] = result["budgetMax"] = 0
        else:
            result["budgetMin"] = result["budgetMax"] = 0

        # PSF
        psf_el = card.select_one("[class*='psf'], [data-testid*='psf']")
        if psf_el:
            psf_text = psf_el.get_text(strip=True)
            psf_nums = re.findall(r"[\d,]+", psf_text.replace(",", ""))
            result["psfAvg"] = int(psf_nums[0]) if psf_nums else 0
        else:
            result["psfAvg"] = 0

        # Bedrooms and size
        beds_el = card.select_one("[data-testid*='beds'], [class*='bed']")
        sqft_el = card.select_one("[data-testid*='area'], [class*='area']")
        result["beds"] = beds_el.get_text(strip=True) if beds_el else ""
        sqft_text = sqft_el.get_text(strip=True) if sqft_el else ""
        sqft_nums = re.findall(r"[\d,]+", sqft_text.replace(",", ""))
        result["sqftMin"] = int(sqft_nums[0]) if sqft_nums else 0
        result["sqftMax"] = int(sqft_nums[-1]) if len(sqft_nums) > 1 else result["sqftMin"]

        # Location / area
        loc_el = card.select_one("[class*='location'], [data-testid*='address']")
        result["area"] = loc_el.get_text(strip=True) if loc_el else ""

        # Listing URL
        link_el = card.select_one("a[href]")
        href = link_el["href"] if link_el else ""
        result["pgListingUrl"] = (
            f"https://www.propertyguru.com.sg{href}" if href.startswith("/") else href
        )

        # District (infer from area text or URL)
        result["district"] = ""
        for d_code in DISTRICT_MAP:
            num = DISTRICT_MAP[d_code]
            if f"D{num}" in result.get("area", "") or f"district={num}" in result.get("pgListingUrl", ""):
                result["district"] = d_code
                break

        result["lastUpdated"] = datetime.date.today().isoformat()

        # Advisor-filled defaults (empty — Chester fills these in Sheet B)
        result["priority"] = "MONITOR"
        result["type"] = "Private Resale"
        result["tenure"] = "99yr"
        result["mrt"] = ""
        result["mrtWalkMins"] = 0
        result["topYear"] = 0
        result["developer"] = ""
        result["totalUnits"] = 0
        result["rent"] = 0
        result["bearCAGR"] = 0.02
        result["baseCAGR"] = 0.04
        result["bullCAGR"] = 0.06
        result["tag"] = ""
        result["thesis"] = ""
        result["catalysts"] = []
        result["risks"] = []
        result["targetUnit"] = ""

    except Exception as e:
        result["_parseError"] = str(e)

    return result


# ═══════════════════════════════════════════════════════════
#  MAIN SCRAPE FUNCTION
# ═══════════════════════════════════════════════════════════
def scrape(params=None):
    """
    Scrape PropertyGuru listings based on params.
    Returns list of listing dicts in ProjectDatabase schema format.
    Falls back to template generation if site is unreachable.
    """
    if params is None:
        params = SEARCH_PARAMS

    results = []
    max_results = params.get("max_results", 20)
    page = 1

    print(f"\n{'='*60}")
    print(f"PropertyGuru Scraper — {datetime.date.today()}")
    print(f"Districts: {params['districts']}")
    print(f"Types: {params['property_types']}")
    print(f"Budget: ${params['budget_min']:,} – ${params['budget_max']:,}")
    print(f"Bedrooms: {params['bedrooms']}BR")
    print(f"{'='*60}\n")

    while len(results) < max_results:
        url = build_search_url(params, page)
        print(f"  Fetching page {page}: {url[:80]}...")

        try:
            resp = SESSION.get(url, timeout=15)
            resp.raise_for_status()
        except requests.exceptions.RequestException as e:
            print(f"\n  [BLOCKED/ERROR] PropertyGuru returned: {e}")
            print("  Switching to template mode — see instructions below.\n")
            return _fallback_template(params)

        soup = BeautifulSoup(resp.text, "html.parser")

        # Try multiple card selectors (PG changes their HTML periodically)
        cards = (
            soup.select("[data-automation-id='listing-card']") or
            soup.select(".listing-card") or
            soup.select("[class*='ListingCard']") or
            soup.select("li.listing-item")
        )

        if not cards:
            print(f"  No listings found on page {page}. Site structure may have changed.")
            break

        for card in cards:
            if len(results) >= max_results:
                break
            listing = parse_listing(card)
            listing["id"] = len(results) + 1
            results.append(listing)
            print(f"  [{listing['id']:02d}] {listing['name']} — ${listing['budgetMin']:,}")

        page += 1
        time.sleep(2)  # polite delay between pages

    print(f"\n  Scraped {len(results)} listings.")
    return results


# ═══════════════════════════════════════════════════════════
#  FALLBACK: TEMPLATE MODE
# ═══════════════════════════════════════════════════════════
def _fallback_template(params):
    """
    If PropertyGuru blocks the scraper, generate a blank template
    with instructions for manual data entry from the site.
    """
    print("\n" + "="*60)
    print("FALLBACK MODE: Manual Entry Template")
    print("="*60)
    print("\nPropertyGuru could not be scraped automatically.")
    print("This happens when:")
    print("  - PropertyGuru has updated their page structure")
    print("  - Your IP has been temporarily rate-limited")
    print("  - You're running without a residential IP")
    print("\nINSTRUCTIONS:")
    print("  1. Open PropertyGuru in your browser")
    print(f"  2. Search for: {params['districts']} | "
          f"{params['property_types']} | "
          f"${params['budget_min']:,}–${params['budget_max']:,} | "
          f"{params['bedrooms']}BR")
    print("  3. For each listing you want to add, fill in the")
    print("     SheetB_Scraped_Results.xlsx template that was generated.")
    print("  4. Key fields: Project Name, District, Budget Range, PSF, Rent estimate")
    print("\nThe template file has been created with blank rows ready to fill.\n")

    # Return blank template rows
    template_rows = []
    for i in range(1, 11):
        template_rows.append({
            "id": i, "name": f"[Project {i} — fill from PropertyGuru]",
            "priority": "MONITOR", "district": params["districts"][0] if params["districts"] else "",
            "area": DISTRICT_AREAS.get(params["districts"][0], "") if params["districts"] else "",
            "type": params["property_types"][0] if params["property_types"] else "Private Resale",
            "tenure": "99yr", "beds": f"{params['bedrooms']}BR",
            "sqftMin": 0, "sqftMax": 0,
            "budgetMin": params["budget_min"], "budgetMax": params["budget_max"],
            "psfAvg": 0, "rent": 0, "mrt": "", "mrtWalkMins": 0,
            "topYear": 0, "developer": "", "totalUnits": 0,
            "bearCAGR": 0.02, "baseCAGR": 0.04, "bullCAGR": 0.06,
            "tag": "", "thesis": "", "catalysts": [], "risks": [], "targetUnit": "",
            "pgListingUrl": "https://www.propertyguru.com.sg/property-for-sale",
            "lastUpdated": datetime.date.today().isoformat(),
        })
    return template_rows


# ═══════════════════════════════════════════════════════════
#  EXPORT TO EXCEL (Sheet B format)
# ═══════════════════════════════════════════════════════════
def export_to_excel(listings, output_path):
    """
    Write listings to an Excel file formatted for Sheet B
    (Phase3_Project_Database_Master.xlsx).
    Includes colour coding and instructions on how to copy into Sheet B.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Results"

    def T(bold=False, size=10, color="000000"):
        return Font(name="Arial", bold=bold, size=size, color=color)
    def F(hex_c):
        return PatternFill("solid", fgColor=hex_c)
    def B():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    BLUE, WHITE, GRAY = "0F4C81", "FFFFFF", "F7F8FA"
    INPUT_BLUE = "0000FF"
    ADVISOR_GRN = "008000"

    # Instructions banner
    ws.merge_cells("A1:Z1")
    ws["A1"] = (
        "SHEET B SCRAPE RESULTS — Review, edit BLUE cells, then copy rows into "
        "Phase3_Project_Database_Master.xlsx  |  "
        f"Scraped: {datetime.date.today()}  |  "
        f"Params: {', '.join(SEARCH_PARAMS['districts'])} | "
        f"${SEARCH_PARAMS['budget_min']:,}–${SEARCH_PARAMS['budget_max']:,}"
    )
    ws["A1"].font = T(bold=True, size=9, color=WHITE)
    ws["A1"].fill = F("0F1923")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # Column definitions matching Sheet B / ProjectDatabase schema
    cols = [
        ("id",          "#",           5,   "number"),
        ("name",        "Project Name", 28,  "text"),
        ("priority",    "Priority",    12,  "text"),
        ("district",    "District",    8,   "text"),
        ("area",        "Area",        18,  "text"),
        ("type",        "Type",        16,  "text"),
        ("tenure",      "Tenure",      10,  "text"),
        ("beds",        "Beds",        8,   "text"),
        ("sqftMin",     "SqFt Min",    10,  "number"),
        ("sqftMax",     "SqFt Max",    10,  "number"),
        ("budgetMin",   "Budget Min",  14,  "currency"),
        ("budgetMax",   "Budget Max",  14,  "currency"),
        ("psfAvg",      "PSF Avg",     10,  "currency"),
        ("rent",        "Rent/mo",     12,  "currency"),
        ("mrt",         "MRT Station", 20,  "text"),
        ("mrtWalkMins", "Walk (min)",  10,  "number"),
        ("topYear",     "TOP Year",    10,  "number"),
        ("developer",   "Developer",   18,  "text"),
        ("totalUnits",  "Total Units", 10,  "number"),
        ("bearCAGR",    "Bear CAGR",   10,  "pct"),
        ("baseCAGR",    "Base CAGR",   10,  "pct"),
        ("bullCAGR",    "Bull CAGR",   10,  "pct"),
        ("tag",         "Tag",         18,  "text"),
        ("thesis",      "Thesis",      40,  "text"),
        ("pgListingUrl","PG URL",      35,  "text"),
        ("lastUpdated", "Updated",     12,  "text"),
    ]

    # Header row
    for col_idx, (key, label, width, fmt) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        c = ws.cell(2, col_idx, label)
        c.font = T(bold=True, size=9, color=WHITE)
        c.fill = F(BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = B()
    ws.row_dimensions[2].height = 24

    # Colour guide
    advisor_cols = {"priority", "mrt", "mrtWalkMins", "rent", "bearCAGR", "baseCAGR", "bullCAGR", "tag", "thesis"}
    scraped_cols = {"name", "district", "area", "budgetMin", "budgetMax", "psfAvg", "sqftMin", "sqftMax",
                    "beds", "topYear", "developer", "totalUnits", "pgListingUrl", "lastUpdated"}

    for row_idx, listing in enumerate(listings, 3):
        ws.row_dimensions[row_idx].height = 18
        for col_idx, (key, label, width, fmt) in enumerate(cols, 1):
            val = listing.get(key, "")
            if isinstance(val, list):
                val = "; ".join(val)

            c = ws.cell(row_idx, col_idx, val if val else None)
            c.border = B()
            c.alignment = Alignment(horizontal="right" if fmt in ("currency","number","pct") else "left",
                                     vertical="center", wrap_text=(fmt == "text"))

            # Color: scraped = black, advisor-fills = blue, URL = gray
            if key == "pgListingUrl":
                c.font = T(size=8, color="888888")
            elif key in advisor_cols:
                c.font = T(size=9, color=INPUT_BLUE)
                c.fill = F("F0F7FF")
            else:
                c.font = T(size=9, color="222222")
                c.fill = F(GRAY if row_idx % 2 == 0 else WHITE)

            # Number formatting
            if fmt == "currency":
                c.number_format = '$#,##0;($#,##0);"-"'
            elif fmt == "pct":
                c.number_format = "0.0%"
            elif fmt == "number":
                c.number_format = "#,##0"

    # Legend
    legend_row = len(listings) + 4
    ws.cell(legend_row, 1, "LEGEND:").font = T(bold=True, size=9)
    ws.cell(legend_row, 2, "Black = scraped from PropertyGuru (verify accuracy)").font = T(size=9, color="222222")
    ws.cell(legend_row+1, 2, "Blue = fill in from your research (rent estimate, MRT, thesis, CAGR, tag)").font = T(size=9, color=INPUT_BLUE)
    ws.cell(legend_row+2, 2, "After filling in blue cells → copy entire rows into Sheet B (Project Database tab)").font = T(size=9, color="008000", bold=True)

    ws.freeze_panes = "C3"
    wb.save(output_path)
    print(f"\n  Saved Excel: {output_path}")


# ═══════════════════════════════════════════════════════════
#  EXPORT TO JSON (for dashboard loading)
# ═══════════════════════════════════════════════════════════
def export_to_json(listings, output_path):
    with open(output_path, "w") as f:
        json.dump({
            "scraped": datetime.date.today().isoformat(),
            "params": SEARCH_PARAMS,
            "count": len(listings),
            "projects": listings,
        }, f, indent=2)
    print(f"  Saved JSON:  {output_path}")


# ═══════════════════════════════════════════════════════════
#  MERGE INTO EXISTING SHEET B (optional helper)
# ═══════════════════════════════════════════════════════════
def merge_into_sheet_b(new_listings, sheet_b_path):
    """
    Append new listings that don't already exist (matched by name)
    into the existing Phase3_Project_Database_Master.xlsx.
    Prints a summary of what was added vs skipped.
    """
    try:
        wb = load_workbook(sheet_b_path)
        ws = wb["Project Database"]
    except Exception as e:
        print(f"  Could not open Sheet B: {e}")
        return

    # Read existing names from Sheet B (row 4 onwards, col 2 = Name)
    existing_names = set()
    for row in ws.iter_rows(min_row=4, max_col=3, values_only=True):
        if row[1]:
            existing_names.add(str(row[1]).strip().lower())

    added, skipped = 0, 0
    max_row = ws.max_row

    for listing in new_listings:
        name = listing.get("name", "").strip()
        if name.lower() in existing_names or not name or "[Project" in name:
            skipped += 1
            continue

        # Append minimal row (advisor fills in the rest)
        next_row = max_row + 1
        ws.cell(next_row, 1, next_row - 3)   # ID
        ws.cell(next_row, 2, name)
        ws.cell(next_row, 3, listing.get("district", ""))
        ws.cell(next_row, 4, listing.get("area", ""))
        ws.cell(next_row, 5, listing.get("type", ""))
        ws.cell(next_row, 10, listing.get("budgetMin", 0))
        ws.cell(next_row, 11, listing.get("budgetMax", 0))
        ws.cell(next_row, 12, listing.get("psfAvg", 0))
        ws.cell(next_row, 14, listing.get("sqftMin", 0))
        ws.cell(next_row, 15, listing.get("sqftMax", 0))
        ws.cell(next_row, 25, listing.get("pgListingUrl", ""))

        existing_names.add(name.lower())
        max_row += 1
        added += 1

    wb.save(sheet_b_path)
    print(f"\n  Sheet B updated: {added} new projects added, {skipped} skipped (already exist or template rows)")


# ═══════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    BASE = "/sessions/stoic-loving-lovelace/mnt/Property matters"
    EXCEL_OUT = f"{BASE}/SheetB_Scraped_Results.xlsx"
    JSON_OUT  = f"{BASE}/SheetB_Scraped_Results.json"
    SHEET_B   = f"{BASE}/Phase3_Project_Database_Master.xlsx"

    # 1. Scrape
    listings = scrape(SEARCH_PARAMS)

    # 2. Export outputs
    export_to_excel(listings, EXCEL_OUT)
    export_to_json(listings, JSON_OUT)

    # 3. Optional: merge into existing Sheet B (comment out if not wanted)
    # merge_into_sheet_b(listings, SHEET_B)

    print(f"\n{'='*60}")
    print("NEXT STEPS:")
    print(f"  1. Open SheetB_Scraped_Results.xlsx")
    print(f"  2. Fill in the BLUE cells (rent, MRT, thesis, CAGR, tag)")
    print(f"  3. Copy completed rows into Phase3_Project_Database_Master.xlsx")
    print(f"  4. Or run: merge_into_sheet_b(listings, SHEET_B) to auto-append")
    print(f"{'='*60}\n")
